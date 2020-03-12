# -*- coding: utf-8 -*-
"""
Created on Wed Aug 09 14:11:14 2017

@author: v-stpurc
"""
from openpyxl import load_workbook

def reversed_string(a_string):
    return a_string[::-1]

def parseFileName(fileName):
    
    str_Length = len(fileName)
    str_rev = reversed_string(fileName)
#    rev_dot_index = str_rev.index('.')
    dot_Index = fileName.index('.')
    file_Extension = fileName[dot_Index + 1:]
    
    rev_Slash_Index = str_rev.index("/")
    slash_Index = str_Length - rev_Slash_Index
    file_Location = fileName[:slash_Index]
    
    file_Name = fileName[slash_Index:dot_Index]
    
    return fileName, file_Name, file_Location, file_Extension

def loadTestSequence(fileName):
    
    PortMap = {}
    TestList = {}
            
    file_Path, file_Name, file_Location, file_Extension = parseFileName(fileName)

    TestList[0] = {}
    TestList[0]['file_Path'] = file_Path 
    TestList[0]['file_Location'] = file_Location 
    TestList[0]['file_Name'] = file_Name 
    TestList[0]['file_Extension'] = file_Extension 


    configWorkBook = load_workbook(filename = fileName, read_only=True, data_only=True)
    
    #Config work book must have tabs matching these names
    #New code will hard code all this stuff, we don't need the functionality David created. That was made for a whole test rack for MTE
    seq_sheet = configWorkBook.get_sheet_by_name(name = 'Test_Sequence') 
    map_sheet = configWorkBook.get_sheet_by_name(name = 'Port_Load_Map') 

    #Look for header cell on Port Map sheet
    #Port Map defines the port names and what individual loads in the Chroma mainframe they are attached to
    for loop in range(1, 30):  #Search for "*header*" cell row
        value = map_sheet.cell(row=loop, column=1).value
        if value == "*header*":
            map_header_row = loop + 1
            break
        
    for loop in range(1, 10):  #Search for columns
        value = map_sheet.cell(row=map_header_row, column=loop).value
    	
        #Frame Num is used for multiple Chroma Mainframes
        if value == "Frame Num":
            map_frame_col = loop 
        #Load Num is the load number within each Chroma Mainframe
        elif value == "Load Num": 
            map_load_col = loop 
        #Console Num is the xbox console under test
        elif value == "Console Num": 
            map_console_col = loop 
        #Port Name is the individual port on the xbox for testing
        elif value == "Port Name":
            map_port_col = loop 
    
    #Can test up to 40 ports, any Chroma mainframe, any xbox 
    #A port is a unique output on a specific xbox and is represented by a 
    #unique combination of Frame Num, Load Num, and Console Num     
    max_number_ports = 40
    port_count = 0                
		
    is_okay = True 
    #This loop counts ports
    #A port is a unique combination of Chroma mainframe, load number, console number
    for loop in range(map_header_row + 1, max_number_ports + map_header_row + 1):
		
        frame     = map_sheet.cell(row=loop, column=map_frame_col).value
        channel   = map_sheet.cell(row=loop, column=map_load_col).value
        console   = map_sheet.cell(row=loop, column=map_console_col).value 
        port      = map_sheet.cell(row=loop, column=map_port_col).value
        if (frame and channel and console and port): # no empty cells in row 
            port_count = port_count + 1 
            ThisPort = {'Frame':frame, 'Channel':channel, 'Console':console, 'Port':port}
            #PortMap is a list of dictionaries each key in PortMap represents a port or a line on the Port_Load_Map tab
            PortMap[port_count] = ThisPort
            #Scan for duplicates went here. For full implementation refer to LUA soucre code by David Furhman. 
        else:
            print "Number of ports " + str(port_count)
            #break will skip over is_okay setting
            break
		
    if port_count == 0: 
        is_okay = False 
    
    #Search for header cell on Test Sequence sheet
    #The test sequence sheet defines the test parameters for each step of the test
    for loop in range(1, 30):  #Search for "*header*" cell row
        value = seq_sheet.cell(row=loop, column=1).value
        if value == "*header*":
            seq_header_row = loop + 1
            break
    
    seq_step_num_col = None 
    seq_console_col = None 
    seq_port_col = None 
    seq_test_id_col = None 
    seq_amps_col = None 
    seq_duration_col = None 
    seq_trigger_col = None 
    seq_trigger_source_col = None 
    seq_volts_open_min_col = None 
    seq_volts_open_max_col = None 
    seq_volts_loaded_min_col = None 
    seq_volts_loaded_max_col = None 
    seq_delay_time_col = None 
    seq_write_csv_col = None 
    seq_excel_command_col = None 
    seq_post_process_col = None 
    seq_slew_rate_col = None 
    seq_rise_slew_rate_col = None 
    seq_fall_slew_rate_col = None 
    seq_pulse_t1_col = None 
    seq_pulse_t1_amps_col = None 
    seq_pulse_t2_col = None 
    seq_pulse_count_col = None 

    #Column names in config spreadsheet must match these although they may not all be required
    #Probably these would never need to change so they could be hard coded
    for loop in range(1, 100):  #Search for columns
        value = seq_sheet.cell(row=seq_header_row, column=loop).value
			
        if value == "Step Num": 
            seq_step_num_col = loop 
        elif value == "Console": 
            seq_console_col = loop 
        elif value == "Port": 
            seq_port_col = loop 
        elif value == "Test ID": 
            seq_test_id_col = loop 
        elif value == "Amps": 
            seq_amps_col = loop 
        elif value == "Duration": 
            seq_duration_col = loop 
        elif value == "Trigger": 
            seq_trigger_col = loop 
        elif value == "Trigger Source": 
            seq_trigger_source_col = loop 
        elif value == "Volts Open Min": 
            seq_volts_open_min_col = loop 
        elif value == "Volts Open Max": 
            seq_volts_open_max_col = loop 
        elif value == "Volts Loaded Min": 
            seq_volts_loaded_min_col = loop 
        elif value == "Volts Loaded Max": 
            seq_volts_loaded_max_col = loop 
        elif value == "Delay Time": 
            seq_delay_time_col = loop 
        elif value == "Raw CSV?": 
            seq_write_csv_col = loop 
        elif value == "Excel New": 
            seq_excel_command_col = loop 
        elif value == "Post Process?": 
            seq_post_process_col = loop 
        elif value == "Slew Rate": 
            seq_slew_rate_col = loop 
        elif value == "Rise Slew Rate": 
            seq_rise_slew_rate_col = loop 
        elif value == "Fall Slew Rate": 
            seq_fall_slew_rate_col = loop 
        elif value == "Pulse T1": 
            seq_pulse_t1_col = loop 
        elif value == "Pulse T1 Amps": 
            seq_pulse_t1_amps_col = loop 
        elif value == "Pulse T2": 
            seq_pulse_t2_col = loop 
        elif value == "Pulse Count": 
            seq_pulse_count_col = loop 

    #This is the max number of steps the code will do
    #This value would represent 100 individual port tests
    #or some lower number of channels combined into group tests
    #For example with five ports grouped 100 would represent 20 individual test runs.
    max_number_tests = 100 
    step_count = 0 
		
    #Go through each step in the test and get values from columns
    for loop in range (seq_header_row + 1, max_number_tests + seq_header_row + 1):
            
        test_id = None
        step_num = seq_sheet.cell(row=loop, column=seq_step_num_col).value 
        #Probably can hard code this to 1 since in our testing we will only ever have one console.
        console = seq_sheet.cell(row=loop, column=seq_console_col).value
        #port is a string
        port = seq_sheet.cell(row=loop, column=seq_port_col).value
        test_id = seq_sheet.cell(row=loop, column=seq_test_id_col).value
        amps = seq_sheet.cell(row=loop, column=seq_amps_col).value
        duration = seq_sheet.cell(row=loop, column=seq_duration_col).value
        #trigger is the time to wait after turning on load before triggering the measurement
        #This is converted to a trigger sample later in the code
        trigger = seq_sheet.cell(row=loop, column=seq_trigger_col).value 

        #Some sheets don't have a trigger source column.
        if seq_trigger_source_col != None:
            trigger_source = seq_sheet.cell(row=loop, column=seq_trigger_source_col).value  
			
        if seq_pulse_t1_col != None: # all pulse columns must be there, so read them all (except pulse_count which is optional) 
            pulse_t1 = seq_sheet.cell(row=loop, column=seq_pulse_t1_col).value   
            pulse_t1_amps = seq_sheet.cell(row=loop, column=seq_pulse_t1_amps_col).value 	
            pulse_t2 = seq_sheet.cell(row=loop, column=seq_pulse_t2_col).value 

        if seq_pulse_count_col != None:
            pulse_count = seq_sheet.cell(row=loop, column=seq_pulse_count_col).value 

        #The following are where specs would be entered to determine pass fail or margin.
        volts_open_min = seq_sheet.cell(row=loop, column=seq_volts_open_min_col).value
        volts_open_max = seq_sheet.cell(row=loop, column=seq_volts_open_max_col).value
        volts_loaded_min = seq_sheet.cell(row=loop, column=seq_volts_loaded_min_col).value
        volts_loaded_max = seq_sheet.cell(row=loop, column=seq_volts_loaded_max_col).value 
        
        delay_time = seq_sheet.cell(row=loop, column=seq_delay_time_col).value 
        write_csv = seq_sheet.cell(row=loop, column=seq_write_csv_col).value 
        excel_command = seq_sheet.cell(row=loop, column=seq_excel_command_col).value 
        post_process = seq_sheet.cell(row=loop, column=seq_post_process_col).value 

        if seq_slew_rate_col != None:
            slew_rate = seq_sheet.cell(row=loop, column=seq_slew_rate_col).value 
            rise_slew_rate = slew_rate
            fall_slew_rate = slew_rate

        if seq_rise_slew_rate_col != None: 
            rise_slew_rate = seq_sheet.cell(row=loop, column=seq_rise_slew_rate_col).value 
			
        if seq_fall_slew_rate_col != None:
            fall_slew_rate = seq_sheet.cell(row=loop, column=seq_fall_slew_rate_col).value 

        #Cell has formulas in it by default. Do these read back as None?
        #No not sure how this worked before. Maybe a LUA thing.
        #Cell below last test ID row must be blank.
        #Here is where we build a complete list reflecting the entirety of the configuration spreadsheet.
        #Beyond this point the spreadsheet is no longer referenced and all the info is in this list.
        if test_id != None: # must have a Test ID, if not, then it's the end of the list 
            step_count = step_count + 1
            #Much of the original LUA code consisted of this kind of check. This error would represent an error
            #in the config spreadsheet and we're already exposed to that issue without having a check
            #Most of this kind of checking has been removed and the onus is on the engineer to avoid errors in the spreadsheet.
            if step_num != step_count: 
                print("Test Sequence import error: Step Num column value and current import count do not match.") 
                print "Step num: " + str(step_num) + " Step Count " + str(step_count)
                print "Test ID : " + str(test_id)
                is_okay = False 
                break
            # Port, Channel, Amps, Duration, Trigger, etc. 
				
            #Create structure holding all information for entire test
            #This structure will have the same number of keys as the number of lines
            #on the Test_Sequence tab regardless of whether these are grouped or individual tests.
            #Test grouping is handled later and is a re-organization of data from this this list.
            TestList[step_count] = {} # instantiate new test entry 
            TestList[step_count]['TestID'] = test_id 
            TestList[step_count]['Console'] = console
            TestList[step_count]['Port'] = port 
            TestList[step_count]['Amps'] = amps 
            TestList[step_count]['Duration'] = duration 
            TestList[step_count]['Trigger'] = trigger 
            TestList[step_count]['TriggerSource'] = trigger_source 
            TestList[step_count]['PulseT1'] = pulse_t1 
            TestList[step_count]['PulseT1Amps'] = pulse_t1_amps  	
            TestList[step_count]['PulseT2'] = pulse_t2  	 
            TestList[step_count]['PulseCount'] = pulse_count  	 
            TestList[step_count]['VoltsOpenMin'] = volts_open_min 
            TestList[step_count]['VoltsOpenMax'] = volts_open_max
            TestList[step_count]['VoltsLoadedMin'] = volts_loaded_min 
            TestList[step_count]['VoltsLoadedMax'] = volts_loaded_max 
            TestList[step_count]['DelayTime'] = delay_time 
            if write_csv == "yes": 
                TestList[step_count]['WriteCsv'] = True 
            else: 
                TestList[step_count]['WriteCsv'] = False 
			
            TestList[step_count]['ExcelCommand'] = excel_command 
            if post_process == "yes": 
                TestList[step_count]['PostProcess'] = True 
            else: 
                TestList[step_count]['PostProcess'] = False 
				
            TestList[step_count]['RiseSlewRate'] = rise_slew_rate 
            TestList[step_count]['FallSlewRate'] = fall_slew_rate 
   
            #Check Test Parameters went here

				
            #Go through Port Map list and check to make sure console number and port name 
            #on test sequence tab have a match on port load map tab.
            # Assign mapped values (Port/Load Map) 
				
            if test_id < 9000: 
                match_found = False 
                #Port count is set in loop above
                for loop2 in range (1, port_count): 
                    ThisPort = PortMap[loop2] 
                    if (ThisPort['Console'] == console) and (ThisPort['Port'] == port): # match found 
                        if match_found: #duplicate, not allowed! (NOTE: this is a backup check to the one done above when importing the Port_Load_Map 
                            #We would only get here if there was a duplicate line on Port Load Map tab
                            print("Internal error: Duplicate port mapping found (backup check).") 
                            is_okay = False 
                            break 
                        else:
                            #TestList is the main list that holds all test parameters
                            #The frame number and load (channel) number are not on the test sequence tab
                            #so copy them from port map list which came from port load map tab.
                            #This section of code should only execute one in each time through the number of tests loop
                            TestList[step_count]['Frame'] = ThisPort['Frame']
                            TestList[step_count]['Channel'] = ThisPort['Channel'] 
                            match_found = True 
					
                if not match_found:
                    print(" **** ERROR: No match found in port/load map") 
                    is_okay = False 
         
        else:   #if testID
            #test ID field was blank
            print "Number of lines on sequence sheet" + str(step_count)
            break
    
    
    return TestList, is_okay
        
 


